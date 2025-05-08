#function to create another sheet in the file and display the timestamp and name of the error along with the inputted preceeding values of all the other parameters

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def newexcel(file, n):
    wb = load_workbook(file)
    wb.create_sheet('Sheet2')
    source = wb['Sheet1']
    target = wb['Sheet2']

    errorlist=['CAN error', 'Out of Buck error','Battery Voltage error','Stack Voltage error', 'Temperature error']
    
    #code segment to input the headings
    max_rowss = source.max_row
    titles = []
    for cell in source[1][1:]:
        titles.append(cell.value)
    target.append(titles)

    #code segment to print the error and preceeding value
    for i in range(2, max_rowss + 1):
        if source.cell(column=9, row=i).value in errorlist:
            firstline=[source.cell(column=2, row=i).value, '','','','','','',source.cell(column=9, row=i).value,'']
            target.append(firstline)
            for j in range(1,n+1):
                row_data = ['']
                for k in range(3, 11):
                    if (i-j)>0:
                        if target.cell(column=2, row=i - j).value !='STACK VOLTAGE':
                            if k==9:
                                row_data.append('')
                            else:
                                row_data.append(source.cell(column=k, row=i - j).value)
                target.append(row_data)

    #colour codes the error and timestamp line 
    max_rowst = target.max_row
    for i in range (2, max_rowst+1):
        if target.cell(column=1, row=i).internal_value!='':
            for j in range(1,10):
                target.cell(row=i, column=j).fill=PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")

    wb.save(file)


