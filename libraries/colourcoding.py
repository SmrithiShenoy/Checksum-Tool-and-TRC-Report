#module to colour code the excel file based on ID values 

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def colourcoding(excel_file_path):
    wb_obj = load_workbook(excel_file_path)
    ws = wb_obj.active
    for i in range (1,11):
        ws.cell(row=1, column=i).fill=PatternFill(start_color="89cff0", end_color="89cff0", fill_type="solid")
    max_rows=ws.max_row
    for i in range (2, max_rows+1):
        if ws.cell(column=1, row=i).internal_value==1458001:
            for j in range(1,11):
                ws.cell(row=i, column=j).fill=PatternFill(start_color="ccffe6", end_color="ccffe6", fill_type="solid")
        elif ws.cell(column=1, row=i).internal_value==1458002:
            for j in range(1,11):
                ws.cell(row=i, column=j).fill=PatternFill(start_color="ffccb3", end_color="ffccb3",fill_type="solid")
                if ws.cell(column=9, row=i).internal_value!="-":
                    ws.cell(row=i, column=j).fill=PatternFill(start_color="ff0000", end_color="ff0000",fill_type="solid")
        elif ws.cell(column=1, row=i).internal_value==1458008:
            for j in range(1,11):
                ws.cell(row=i, column=j).fill=PatternFill(start_color="e699ff", end_color="e699ff",fill_type="solid")
        elif ws.cell(column=1, row=i).internal_value==1804501:
            for j in range (1,11):
                ws.cell(row=i, column=j).fill=PatternFill(start_color="fffacd", end_color="e699ff",fill_type="solid")

    wb_obj.save(excel_file_path)
    print("Color coding applied and Excel file saved.")